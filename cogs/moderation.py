"""
Moderation — server utility commands for owners and admins.

Commands:
  /role                — Assign/toggle a safe-listed role on a member.
  /bulk-role           — Modal: assign a role to many users at once.
  /new-role            — Create a new role with a name and optional hex color.
  /rename-channel      — Rename a channel (protects log and announcement channels).
  /toggle-channel      — Toggle view_channel permission for a role on a channel.
  /sync-channels       — Sync all channels in a category to match category permissions.
  /restrict-channel    — Restrict or unrestrict a user to a specific channel or category.
  /lock-threads        — Lock and archive all threads in a channel.
  /export              — Export all members with their roles to a CSV file.
  /export-category     — Export all messages per channel in a category to a zip of CSVs.
  /export-permissions  — Export all roles and permissions to a colour-coded Excel file.
  /list-overrides      — List all channels with user-specific permission overrides.

All commands require server owner or bot owner access.
"""

import csv
import io
import zipfile
import discord
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from discord.ext import commands
from discord.commands import Option
import db_handler
import permissions
import two_factor_helper
import logger


# Permissions that make a role too dangerous to assign via /role or /bulk-role
DANGEROUS_PERMISSIONS = [
    "administrator", "manage_guild", "manage_roles", "manage_channels",
    "ban_members", "kick_members", "manage_webhooks", "view_audit_log",
    "manage_expressions", "manage_threads", "mention_everyone", "moderate_members",
]

# ---------------------------------------------------------------------------
# export-permissions — Excel styling constants
# ---------------------------------------------------------------------------

_GREEN_FILL  = PatternFill(fill_type="solid", fgColor="C6EFCE")   # allowed
_RED_FILL    = PatternFill(fill_type="solid", fgColor="FFC7CE")   # denied
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="23272A")   # Discord dark
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BOLD        = Font(bold=True)
_CENTER      = Alignment(horizontal="center", vertical="center")

_CHANNEL_TYPE = {
    "TextChannel":  "Text",
    "VoiceChannel": "Voice",
    "StageChannel": "Stage",
    "ForumChannel": "Forum",
    "NewsChannel":  "News",
}

# Every Discord permission with a human-readable label (ordered by category)
_PERM_LABELS = {
    # General server
    "administrator":            "Administrator",
    "manage_guild":             "Manage Server",
    "manage_roles":             "Manage Roles",
    "manage_channels":          "Manage Channels",
    "manage_webhooks":          "Manage Webhooks",
    "manage_expressions":       "Manage Expressions",
    "manage_events":            "Manage Events",
    "manage_threads":           "Manage Threads",
    "manage_nicknames":         "Manage Nicknames",
    "kick_members":             "Kick Members",
    "ban_members":              "Ban Members",
    "moderate_members":         "Timeout Members",
    "view_audit_log":           "View Audit Log",
    "view_guild_insights":      "View Server Insights",
    "create_instant_invite":    "Create Invites",
    "change_nickname":          "Change Nickname",
    "mention_everyone":         "Mention @everyone",
    # Text
    "view_channel":             "View Channels",
    "send_messages":            "Send Messages",
    "send_messages_in_threads": "Send in Threads",
    "create_public_threads":    "Create Public Threads",
    "create_private_threads":   "Create Private Threads",
    "manage_messages":          "Manage Messages",
    "embed_links":              "Embed Links",
    "attach_files":             "Attach Files",
    "add_reactions":            "Add Reactions",
    "use_external_emojis":      "Use External Emojis",
    "use_external_stickers":    "Use External Stickers",
    "read_message_history":     "Read Message History",
    "send_tts_messages":        "Send TTS Messages",
    "use_application_commands": "Use Slash Commands",
    "send_voice_messages":      "Send Voice Messages",
    "request_to_speak":         "Request to Speak (Stage)",
    # Voice
    "connect":                  "Connect (Voice)",
    "speak":                    "Speak (Voice)",
    "stream":                   "Video / Go Live",
    "use_voice_activation":     "Use Voice Activity",
    "priority_speaker":         "Priority Speaker",
    "mute_members":             "Mute Members (Voice)",
    "deafen_members":           "Deafen Members (Voice)",
    "move_members":             "Move Members (Voice)",
    "use_embedded_activities":  "Use Activities",
    "use_soundboard":           "Use Soundboard",
    "use_external_sounds":      "Use External Sounds",
}


def role_has_dangerous_perms(role: discord.Role) -> bool:
    return any(getattr(role.permissions, p, False) for p in DANGEROUS_PERMISSIONS)


# ---------------------------------------------------------------------------
# Bulk-role modal
# ---------------------------------------------------------------------------

class BulkRoleModal(discord.ui.Modal):
    def __init__(self, bot, guild: discord.Guild):
        super().__init__(title="Bulk Role Assignment")
        self.bot = bot
        self.guild = guild
        self.add_item(discord.ui.InputText(
            label="User IDs (one per line or comma-separated)",
            style=discord.InputTextStyle.paragraph,
            placeholder="123456789012345678\n987654321098765432",
            required=True,
            max_length=4000,
        ))
        self.add_item(discord.ui.InputText(
            label="Role ID to assign",
            style=discord.InputTextStyle.short,
            placeholder="123456789012345678",
            required=True,
            max_length=20,
        ))

    async def callback(self, interaction: discord.Interaction):
        raw_ids = self.children[0].value
        role_id_str = self.children[1].value.strip()

        try:
            role_id = int(role_id_str)
        except ValueError:
            await interaction.response.send_message("Invalid role ID.", ephemeral=True)
            return

        role = self.guild.get_role(role_id)
        if role is None:
            await interaction.response.send_message("Role not found in this server.", ephemeral=True)
            return

        if role_has_dangerous_perms(role):
            await interaction.response.send_message(
                f"**{role.name}** has dangerous permissions and cannot be bulk-assigned.",
                ephemeral=True,
            )
            return

        if not db_handler.is_safe_role(self.bot.CONN, self.guild.id, role_id):
            await interaction.response.send_message(
                f"**{role.name}** is not on the safe-role whitelist. "
                "Add it first — ask an admin.",
                ephemeral=True,
            )
            return

        raw_list = raw_ids.replace(",", "\n").splitlines()
        user_ids = [int(r.strip()) for r in raw_list if r.strip().isdigit()]

        if not user_ids:
            await interaction.response.send_message("No valid user IDs found.", ephemeral=True)
            return

        await interaction.response.defer(ephemeral=True)

        success, failed = [], []
        for uid in user_ids:
            try:
                member = self.guild.get_member(uid) or await self.guild.fetch_member(uid)
                await member.add_roles(role, reason=f"Bulk role by {interaction.user}")
                success.append(uid)
            except (discord.NotFound, discord.Forbidden, discord.HTTPException):
                failed.append(uid)

        await interaction.followup.send(
            f"Assigned **{role.name}** to **{len(success)}** members."
            + (f" {len(failed)} failed." if failed else ""),
            ephemeral=True,
        )

        await logger.log_action(
            self.bot, self.guild, "Bulk Role Assignment", interaction.user,
            details={
                "Role": role.name,
                "Assigned": str(len(success)),
                "Failed": str(len(failed)),
                **({"Failed IDs": ", ".join(str(i) for i in failed[:10])} if failed else {}),
            },
            level='info'
        )


# ---------------------------------------------------------------------------
# Cog
# ---------------------------------------------------------------------------

class Moderation(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    def _check_owner(self, ctx) -> tuple[bool, str]:
        return permissions.check(self.bot, ctx, 'owner')

    # ------------------------------------------------------------------
    # /role
    # ------------------------------------------------------------------

    @commands.guild_only()
    @commands.slash_command(description="[Owner] Assign or remove a safe-listed role from a member. Requires 2FA.")
    async def role(self, ctx: discord.ApplicationContext,
                   member: Option(discord.Member, "Target member"),
                   role: Option(discord.Role, "Role to assign or remove"),
                   code: Option(int, "Your 6-digit 2FA code", required=True)):
        allowed, err = self._check_owner(ctx)
        if not allowed:
            await ctx.respond(err, ephemeral=True)
            return

        ok, err2 = permissions.guild_required(self.bot, ctx)
        if not ok:
            await ctx.respond(err2, ephemeral=True)
            return

        if not two_factor_helper.verify_code(self.bot.CONN, ctx.author.id, code):
            await ctx.respond("Incorrect 2FA code.", ephemeral=True)
            return

        if role_has_dangerous_perms(role):
            await ctx.respond(
                f"**{role.name}** has dangerous permissions and cannot be assigned via this command.",
                ephemeral=True,
            )
            return

        if not db_handler.is_safe_role(self.bot.CONN, ctx.guild.id, role.id):
            await ctx.respond(
                f"**{role.name}** is not on the safe-role whitelist. "
                "Ask an admin to add it first.",
                ephemeral=True,
            )
            return

        if role >= ctx.guild.me.top_role:
            await ctx.respond("That role is above my highest role. I cannot manage it.", ephemeral=True)
            return

        if role in member.roles:
            await member.remove_roles(role, reason=f"/role by {ctx.author}")
            action = "removed from"
        else:
            await member.add_roles(role, reason=f"/role by {ctx.author}")
            action = "assigned to"

        await ctx.respond(f"**{role.name}** {action} {member.mention}.", ephemeral=True)
        await logger.log_action(
            self.bot, ctx.guild, "Role Updated", ctx.author,
            details={"Role": role.name, "Member": str(member), "Action": action},
            level='info'
        )

    # ------------------------------------------------------------------
    # /bulk-role
    # ------------------------------------------------------------------

    @commands.guild_only()
    @commands.slash_command(
        name="bulk-role",
        description="[Owner] Open a modal to assign a role to many users at once. Requires 2FA."
    )
    async def bulk_role(self, ctx: discord.ApplicationContext,
                        code: Option(int, "Your 6-digit 2FA code", required=True)):
        allowed, err = self._check_owner(ctx)
        if not allowed:
            await ctx.respond(err, ephemeral=True)
            return

        ok, err2 = permissions.guild_required(self.bot, ctx)
        if not ok:
            await ctx.respond(err2, ephemeral=True)
            return

        if not two_factor_helper.verify_code(self.bot.CONN, ctx.author.id, code):
            await ctx.respond("Incorrect 2FA code.", ephemeral=True)
            return

        await ctx.send_modal(BulkRoleModal(self.bot, ctx.guild))

    # ------------------------------------------------------------------
    # /new-role
    # ------------------------------------------------------------------

    @commands.guild_only()
    @commands.slash_command(
        name="new-role",
        description="[Owner] Create a new role with a specified name and optional color. Requires 2FA."
    )
    async def new_role(self, ctx: discord.ApplicationContext,
                       name: Option(str, "Role name", max_length=100, required=True),
                       code: Option(int, "Your 6-digit 2FA code", required=True),
                       color: Option(str, "Hex color (e.g. #ff0000). Omit for default.", required=False)):
        allowed, err = self._check_owner(ctx)
        if not allowed:
            await ctx.respond(err, ephemeral=True)
            return

        if not two_factor_helper.verify_code(self.bot.CONN, ctx.author.id, code):
            await ctx.respond("Incorrect 2FA code.", ephemeral=True)
            return

        color_val = discord.Color.default()
        if color:
            try:
                color_val = discord.Color(int(color.strip("#"), 16))
            except ValueError:
                await ctx.respond("Invalid hex color. Use format #rrggbb, e.g. `#ff0000`.", ephemeral=True)
                return

        try:
            new = await ctx.guild.create_role(
                name=name, color=color_val, reason=f"/new-role by {ctx.author}"
            )
        except discord.Forbidden:
            await ctx.respond("Missing permissions to create roles.", ephemeral=True)
            return

        await ctx.respond(f"Created role {new.mention}.", ephemeral=True)
        await logger.log_action(
            self.bot, ctx.guild, "Role Created", ctx.author,
            details={"Role": new.name, "Color": str(color_val)},
            level='info'
        )

    # ------------------------------------------------------------------
    # /rename-channel
    # ------------------------------------------------------------------

    @commands.guild_only()
    @commands.slash_command(
        name="rename-channel",
        description="[Owner] Rename a channel. Log and announcement channels are protected. Requires 2FA."
    )
    async def rename_channel(self, ctx: discord.ApplicationContext,
                              channel: Option(discord.TextChannel, "Channel to rename"),
                              new_name: Option(str, "New name (hyphens for spaces)", max_length=100),
                              code: Option(int, "Your 6-digit 2FA code", required=True)):
        allowed, err = self._check_owner(ctx)
        if not allowed:
            await ctx.respond(err, ephemeral=True)
            return

        ok, err2 = permissions.guild_required(self.bot, ctx)
        if not ok:
            await ctx.respond(err2, ephemeral=True)
            return

        if not two_factor_helper.verify_code(self.bot.CONN, ctx.author.id, code):
            await ctx.respond("Incorrect 2FA code.", ephemeral=True)
            return

        if channel.id in db_handler.get_channels(self.bot.CONN, ctx.guild.id):
            await ctx.respond("That is a protected announcement channel.", ephemeral=True)
            return

        log_id = db_handler.get_log_channel(self.bot.CONN, ctx.guild.id)
        if channel.id == log_id:
            await ctx.respond("The log channel cannot be renamed via this command.", ephemeral=True)
            return

        old_name = channel.name
        clean_name = new_name.lower().replace(" ", "-")

        try:
            await channel.edit(name=clean_name, reason=f"/rename-channel by {ctx.author}")
        except discord.Forbidden:
            await ctx.respond("Missing permissions to rename that channel.", ephemeral=True)
            return

        await ctx.respond(f"Renamed **#{old_name}** → **#{clean_name}**.", ephemeral=True)
        await logger.log_action(
            self.bot, ctx.guild, "Channel Renamed", ctx.author,
            details={"Before": f"#{old_name}", "After": f"#{clean_name}"},
            level='info'
        )

    # ------------------------------------------------------------------
    # /toggle-channel
    # ------------------------------------------------------------------

    @commands.guild_only()
    @commands.slash_command(
        name="toggle-channel",
        description="[Owner] Toggle view permission for a role on a channel. Requires 2FA."
    )
    async def toggle_channel(self, ctx: discord.ApplicationContext,
                              channel: Option(discord.abc.GuildChannel, "Target channel"),
                              role: Option(discord.Role, "Role to toggle"),
                              code: Option(int, "Your 6-digit 2FA code", required=True)):
        allowed, err = self._check_owner(ctx)
        if not allowed:
            await ctx.respond(err, ephemeral=True)
            return

        if not two_factor_helper.verify_code(self.bot.CONN, ctx.author.id, code):
            await ctx.respond("Incorrect 2FA code.", ephemeral=True)
            return

        current = channel.overwrites_for(role)
        currently_hidden = current.view_channel is False

        if currently_hidden:
            current.view_channel = None
            old_str, new_str = "Hidden", "Visible"
            log_action = "Unhid channel for role"
        else:
            current.view_channel = False
            old_str, new_str = "Visible", "Hidden"
            log_action = "Hid channel from role"

        try:
            await channel.set_permissions(role, overwrite=current, reason=f"/toggle-channel by {ctx.author}")
        except discord.Forbidden:
            await ctx.respond("Missing permissions to edit that channel.", ephemeral=True)
            return

        await ctx.respond(
            f"{channel.mention} for {role.mention}: **{old_str}** → **{new_str}**.",
            ephemeral=True,
        )
        await logger.log_action(
            self.bot, ctx.guild, "Channel Visibility Toggled", ctx.author,
            details={
                "Channel": channel.name,
                "Role": role.name,
                "Previous": old_str,
                "New State": new_str,
            },
            level='info'
        )

    # ------------------------------------------------------------------
    # /sync-channels
    # ------------------------------------------------------------------

    @commands.guild_only()
    @commands.slash_command(
        name="sync-channels",
        description="[Owner] Sync all channels in a category to match the category's permissions. Requires 2FA."
    )
    async def sync_channels(self, ctx: discord.ApplicationContext,
                             category: Option(discord.CategoryChannel, "Category to sync", required=True),
                             code: Option(int, "Your 6-digit 2FA code", required=True)):
        allowed, err = self._check_owner(ctx)
        if not allowed:
            await ctx.respond(err, ephemeral=True)
            return

        if not two_factor_helper.verify_code(self.bot.CONN, ctx.author.id, code):
            await ctx.respond("Incorrect 2FA code.", ephemeral=True)
            return

        await ctx.defer(ephemeral=True)
        synced, errors = 0, 0
        for ch in category.channels:
            try:
                await ch.edit(sync_permissions=True, reason=f"/sync-channels by {ctx.author}")
                synced += 1
            except (discord.Forbidden, discord.HTTPException):
                errors += 1

        await ctx.respond(f"Synced **{synced}** channels. {errors} error(s).", ephemeral=True)
        await logger.log_action(
            self.bot, ctx.guild, "Channels Synced", ctx.author,
            details={"Category": category.name, "Synced": str(synced), "Errors": str(errors)},
            level='info'
        )

    # ------------------------------------------------------------------
    # /restrict-channel
    # ------------------------------------------------------------------

    @commands.guild_only()
    @commands.slash_command(
        name="restrict-channel",
        description="[Owner] Restrict or unrestrict a user to a specific channel or category. Requires 2FA."
    )
    async def restrict_channel(self, ctx: discord.ApplicationContext,
                                member: Option(discord.Member, "Member to restrict/unrestrict"),
                                action: Option(str, "Action", choices=["restrict", "unrestrict"], required=True),
                                channel: Option(discord.abc.GuildChannel, "Channel or category target", required=True),
                                code: Option(int, "Your 6-digit 2FA code", required=True)):
        allowed, err = self._check_owner(ctx)
        if not allowed:
            await ctx.respond(err, ephemeral=True)
            return

        if not two_factor_helper.verify_code(self.bot.CONN, ctx.author.id, code):
            await ctx.respond("Incorrect 2FA code.", ephemeral=True)
            return

        await ctx.defer(ephemeral=True)
        errors = 0

        if action == "restrict":
            # Deny view_channel in all categories for this user
            for cat in ctx.guild.categories:
                try:
                    ow = cat.overwrites_for(member)
                    ow.view_channel = False
                    await cat.set_permissions(member, overwrite=ow, reason=f"/restrict-channel by {ctx.author}")
                except (discord.Forbidden, discord.HTTPException):
                    errors += 1

            # Allow access in the target channel or category
            allow_ow = discord.PermissionOverwrite(view_channel=True, send_messages=True)
            try:
                await channel.set_permissions(member, overwrite=allow_ow, reason=f"/restrict-channel by {ctx.author}")
            except (discord.Forbidden, discord.HTTPException):
                errors += 1

            await ctx.respond(
                f"{member.mention} restricted to {channel.mention}. ({errors} error(s))",
                ephemeral=True
            )

        elif action == "unrestrict":
            # Remove all user-specific permission overwrites across all channels
            for ch in ctx.guild.channels:
                if member in ch.overwrites:
                    try:
                        await ch.set_permissions(member, overwrite=None, reason=f"/restrict-channel unrestrict by {ctx.author}")
                    except (discord.Forbidden, discord.HTTPException):
                        errors += 1

            await ctx.respond(
                f"{member.mention} unrestricted. All overrides removed. ({errors} error(s))",
                ephemeral=True
            )

        await logger.log_action(
            self.bot, ctx.guild, f"Channel Restriction {action.title()}d", ctx.author,
            details={
                "Member": f"{member} ({member.id})",
                "Target": channel.name,
                "Errors": str(errors),
            },
            level='info'
        )

    # ------------------------------------------------------------------
    # /lock-threads
    # ------------------------------------------------------------------

    @commands.guild_only()
    @commands.slash_command(
        name="lock-threads",
        description="[Owner] Lock and archive all threads in a channel, including old archived ones. Requires 2FA."
    )
    async def lock_threads(self, ctx: discord.ApplicationContext,
                           code: Option(int, "Your 6-digit 2FA code", required=True),
                           channel: Option(discord.TextChannel, "Channel with threads to lock", required=False)):
        allowed, err = self._check_owner(ctx)
        if not allowed:
            await ctx.respond(err, ephemeral=True)
            return

        if not two_factor_helper.verify_code(self.bot.CONN, ctx.author.id, code):
            await ctx.respond("Incorrect 2FA code.", ephemeral=True)
            return

        target = channel or ctx.channel
        if not isinstance(target, discord.TextChannel):
            await ctx.respond("Target must be a text channel.", ephemeral=True)
            return

        await ctx.defer(ephemeral=True)

        all_threads = list(target.threads)
        try:
            async for archived in target.archived_threads(limit=None):
                all_threads.append(archived)
        except (discord.Forbidden, discord.HTTPException):
            pass

        locked, errors = 0, 0
        for thread in all_threads:
            try:
                await thread.edit(
                    archived=True,
                    locked=True,
                    reason=f"/lock-threads by {ctx.author}"
                )
                locked += 1
            except (discord.Forbidden, discord.HTTPException):
                errors += 1

        await ctx.respond(
            f"Locked and archived **{locked}** thread(s) in {target.mention}. {errors} error(s).",
            ephemeral=True
        )
        await logger.log_action(
            self.bot, ctx.guild, "Threads Locked", ctx.author,
            details={"Channel": target.name, "Locked": str(locked), "Errors": str(errors)},
            level='info'
        )

    # ------------------------------------------------------------------
    # /export
    # ------------------------------------------------------------------

    @commands.guild_only()
    @commands.slash_command(
        description="[Owner] Export all server members with their roles to a CSV file. Requires 2FA."
    )
    async def export(self, ctx: discord.ApplicationContext,
                     code: Option(int, "Your 6-digit 2FA code", required=True)):
        allowed, err = self._check_owner(ctx)
        if not allowed:
            await ctx.respond(err, ephemeral=True)
            return

        if not two_factor_helper.verify_code(self.bot.CONN, ctx.author.id, code):
            await ctx.respond("Incorrect 2FA code.", ephemeral=True)
            return

        await ctx.defer(ephemeral=True)

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["User ID", "Username", "Display Name", "Roles"])

        for member in ctx.guild.members:
            roles = [r.name for r in member.roles if r.name != "@everyone"]
            writer.writerow([member.id, str(member), member.display_name, ", ".join(roles)])

        output.seek(0)
        csv_bytes = io.BytesIO(output.getvalue().encode("utf-8"))
        file = discord.File(csv_bytes, filename=f"members_{ctx.guild.name}.csv")

        await ctx.respond("Member export:", file=file, ephemeral=True)
        await logger.log_action(
            self.bot, ctx.guild, "Member List Exported", ctx.author,
            details={"Members": str(ctx.guild.member_count)},
            level='info'
        )

    # ------------------------------------------------------------------
    # /export-category
    # ------------------------------------------------------------------

    @commands.guild_only()
    @commands.slash_command(
        name="export-category",
        description="[Owner] Export all messages from each channel in a category as a ZIP of CSVs. Requires 2FA."
    )
    async def export_category(self, ctx: discord.ApplicationContext,
                               category: Option(discord.CategoryChannel, "Category to export", required=True),
                               code: Option(int, "Your 6-digit 2FA code", required=True)):
        allowed, err = self._check_owner(ctx)
        if not allowed:
            await ctx.respond(err, ephemeral=True)
            return

        if not two_factor_helper.verify_code(self.bot.CONN, ctx.author.id, code):
            await ctx.respond("Incorrect 2FA code.", ephemeral=True)
            return

        text_channels = [ch for ch in category.channels if isinstance(ch, discord.TextChannel)]
        if not text_channels:
            await ctx.respond("No text channels found in that category.", ephemeral=True)
            return

        await ctx.defer(ephemeral=True)

        zip_buffer = io.BytesIO()
        total_messages = 0

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for ch in text_channels:
                csv_output = io.StringIO()
                writer = csv.writer(csv_output)
                writer.writerow(["Message ID", "Timestamp", "Author ID", "Author", "Content"])
                count = 0
                try:
                    async for msg in ch.history(limit=1000, oldest_first=True):
                        writer.writerow([
                            msg.id,
                            msg.created_at.isoformat(),
                            msg.author.id,
                            str(msg.author),
                            msg.content[:500],
                        ])
                        count += 1
                except (discord.Forbidden, discord.HTTPException):
                    writer.writerow(["ERROR", "Could not access channel history", "", "", ""])
                total_messages += count
                zf.writestr(f"{ch.name}.csv", csv_output.getvalue())

        zip_buffer.seek(0)
        file = discord.File(zip_buffer, filename=f"category_{category.name}.zip")

        await ctx.respond(
            f"Exported {len(text_channels)} channel(s), {total_messages} message(s).",
            file=file,
            ephemeral=True
        )
        await logger.log_action(
            self.bot, ctx.guild, "Category Exported", ctx.author,
            details={
                "Category": category.name,
                "Channels": str(len(text_channels)),
                "Messages": str(total_messages),
            },
            level='info'
        )

    # ------------------------------------------------------------------
    # /export-permissions
    # ------------------------------------------------------------------

    @commands.guild_only()
    @commands.slash_command(
        name="export-permissions",
        description="[Owner] Export all roles and their permissions to a colour-coded Excel file. Requires 2FA."
    )
    async def export_permissions(self, ctx: discord.ApplicationContext,
                                 code: Option(int, "Your 6-digit 2FA code", required=True)):
        allowed, err = self._check_owner(ctx)
        if not allowed:
            await ctx.respond(err, ephemeral=True)
            return

        if not two_factor_helper.verify_code(self.bot.CONN, ctx.author.id, code):
            await ctx.respond("Incorrect 2FA code.", ephemeral=True)
            return

        await ctx.defer(ephemeral=True)

        guild = ctx.guild
        # Highest role first so Administrator appears near the left
        roles = list(reversed(guild.roles))

        wb = openpyxl.Workbook()

        # ── Sheet 1: Server-Level Permissions ──────────────────────────────
        ws1 = wb.active
        ws1.title = "Server Permissions"
        ws1.freeze_panes = "B2"  # Keep first row + column locked while scrolling

        # Corner label
        corner = ws1.cell(1, 1, "Permission")
        corner.font = _HEADER_FONT
        corner.fill = _HEADER_FILL
        corner.alignment = _CENTER

        # Role name headers (one column per role)
        for col_idx, role in enumerate(roles, start=2):
            cell = ws1.cell(1, col_idx, role.name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER

        # One row per permission
        for row_idx, (perm_key, perm_label) in enumerate(_PERM_LABELS.items(), start=2):
            ws1.cell(row_idx, 1, perm_label).font = _BOLD
            for col_idx, role in enumerate(roles, start=2):
                # Administrator implicitly grants every permission
                if role.permissions.administrator:
                    has_perm = True
                else:
                    has_perm = getattr(role.permissions, perm_key, False)
                cell = ws1.cell(row_idx, col_idx, "✅" if has_perm else "❌")
                cell.alignment = _CENTER
                cell.fill = _GREEN_FILL if has_perm else _RED_FILL

        # Column widths
        ws1.column_dimensions["A"].width = 28
        for col_idx, role in enumerate(roles, start=2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws1.column_dimensions[col_letter].width = max(len(role.name) + 2, 12)

        # ── Sheet 2: Category Overrides ────────────────────────────────────
        ws2 = wb.create_sheet("Category Overrides")
        for col_idx, header in enumerate(["Category", "Role", "Permission", "Override"], start=1):
            cell = ws2.cell(1, col_idx, header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER

        has_cat_data = False
        for category in guild.categories:
            for target, overwrite in category.overwrites.items():
                if not isinstance(target, discord.Role):
                    continue
                allow_perms, deny_perms = overwrite.pair()
                for perm_key, perm_label in _PERM_LABELS.items():
                    is_allow = getattr(allow_perms, perm_key, False)
                    is_deny  = getattr(deny_perms,  perm_key, False)
                    if not is_allow and not is_deny:
                        continue  # ➖ Inherited — skip to keep sheet concise
                    has_cat_data = True
                    override_str = "✅ Allow" if is_allow else "❌ Deny"
                    row = ws2.max_row + 1
                    ws2.cell(row, 1, category.name)
                    ws2.cell(row, 2, target.name)
                    ws2.cell(row, 3, perm_label)
                    ov_cell = ws2.cell(row, 4, override_str)
                    ov_cell.fill = _GREEN_FILL if is_allow else _RED_FILL
                    ov_cell.alignment = _CENTER

        if not has_cat_data:
            ws2.cell(2, 1, "No role overrides found on any category.")

        for col_letter, width in zip("ABCD", [25, 22, 28, 12]):
            ws2.column_dimensions[col_letter].width = width

        # ── Sheet 3: Channel Overrides ─────────────────────────────────────
        ws3 = wb.create_sheet("Channel Overrides")
        for col_idx, header in enumerate(
            ["Channel", "Type", "Category", "Role", "Permission", "Override"], start=1
        ):
            cell = ws3.cell(1, col_idx, header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER

        has_ch_data = False
        sorted_channels = sorted(
            (c for c in guild.channels if not isinstance(c, discord.CategoryChannel)),
            key=lambda c: (c.category.position if c.category else -1, c.position)
        )
        for channel in sorted_channels:
            category_name = channel.category.name if channel.category else "— No Category —"
            ch_type = _CHANNEL_TYPE.get(type(channel).__name__, "Channel")
            for target, overwrite in channel.overwrites.items():
                if not isinstance(target, discord.Role):
                    continue
                allow_perms, deny_perms = overwrite.pair()
                for perm_key, perm_label in _PERM_LABELS.items():
                    is_allow = getattr(allow_perms, perm_key, False)
                    is_deny  = getattr(deny_perms,  perm_key, False)
                    if not is_allow and not is_deny:
                        continue
                    has_ch_data = True
                    override_str = "✅ Allow" if is_allow else "❌ Deny"
                    row = ws3.max_row + 1
                    ws3.cell(row, 1, channel.name)
                    ws3.cell(row, 2, ch_type)
                    ws3.cell(row, 3, category_name)
                    ws3.cell(row, 4, target.name)
                    ws3.cell(row, 5, perm_label)
                    ov_cell = ws3.cell(row, 6, override_str)
                    ov_cell.fill = _GREEN_FILL if is_allow else _RED_FILL
                    ov_cell.alignment = _CENTER

        if not has_ch_data:
            ws3.cell(2, 1, "No role overrides found on any channel.")

        for col_letter, width in zip("ABCDEF", [25, 10, 22, 22, 28, 12]):
            ws3.column_dimensions[col_letter].width = width

        # ── Send file ──────────────────────────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        non_cat_channels = [c for c in guild.channels if not isinstance(c, discord.CategoryChannel)]
        file = discord.File(buffer, filename=f"permissions_{guild.name}.xlsx")
        await ctx.respond(
            f"Permission export ready — **{len(roles)}** roles · "
            f"**{len(guild.categories)}** categories · "
            f"**{len(non_cat_channels)}** channels.",
            file=file,
            ephemeral=True,
        )

        await logger.log_action(
            self.bot, ctx.guild, "Permissions Exported", ctx.author,
            details={
                "Roles": str(len(roles)),
                "Categories": str(len(guild.categories)),
                "Channels": str(len(non_cat_channels)),
            },
            level='info'
        )

    # ------------------------------------------------------------------
    # /list-overrides
    # ------------------------------------------------------------------

    @commands.guild_only()
    @commands.slash_command(
        name="list-overrides",
        description="[Owner] List all channels with user-specific permission overrides. Requires 2FA."
    )
    async def list_overrides(self, ctx: discord.ApplicationContext,
                             code: Option(int, "Your 6-digit 2FA code", required=True)):
        allowed, err = self._check_owner(ctx)
        if not allowed:
            await ctx.respond(err, ephemeral=True)
            return

        if not two_factor_helper.verify_code(self.bot.CONN, ctx.author.id, code):
            await ctx.respond("Incorrect 2FA code.", ephemeral=True)
            return

        await ctx.defer(ephemeral=True)

        results = []
        for channel in ctx.guild.channels:
            user_overrides = [
                (target, ow)
                for target, ow in channel.overwrites.items()
                if isinstance(target, discord.Member)
            ]
            if user_overrides:
                results.append((channel, user_overrides))

        if not results:
            await ctx.respond("No user-specific permission overrides found.", ephemeral=True)
            return

        embed = discord.Embed(
            title="User Permission Overrides",
            color=0xf39c12,
            description=f"{sum(len(ow) for _, ow in results)} override(s) across {len(results)} channel(s)."
        )

        for channel, overrides in results[:15]:
            lines = []
            for member, ow in overrides[:5]:
                allow_bits, deny_bits = ow.pair()
                lines.append(
                    f"• {member.mention}: "
                    f"allow `{allow_bits.value}` / deny `{deny_bits.value}`"
                )
            embed.add_field(
                name=f"#{channel.name}",
                value="\n".join(lines) or "none",
                inline=False
            )

        if len(results) > 15:
            embed.set_footer(text=f"Showing first 15 of {len(results)} channels.")

        await ctx.respond(embed=embed, ephemeral=True)


def setup(bot):
    bot.add_cog(Moderation(bot))
